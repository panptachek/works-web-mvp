#!/usr/bin/env python3
import os
import re
import hashlib
from pathlib import Path
from decimal import Decimal
import openpyxl
import psycopg2

DB = dict(
    dbname=os.getenv('DB_NAME', 'works_db_v2'),
    user=os.getenv('DB_USER', 'works_user'),
    password=os.getenv('DB_PASSWORD', ''),
    host=os.getenv('DB_HOST', '127.0.0.1'),
    port=int(os.getenv('DB_PORT', '5433')),
)
REPORT_PATH = Path('/home/aboba/.openclaw/workspace/input_report.xlsx')
SOURCE_TYPE = 'xlsx_report'
SECTION_SHEETS = [f'Уч. {i}' for i in range(1, 9)]
COMMON_SHEET = 'Общ'

PK_RANGE_RE = re.compile(r'ПК\s*(\d+[+,\.]?\d*)\s*[-–]\s*ПК?\s*(\d+[+,\.]?\d*)', re.I)
PK_ANY_RE = re.compile(r'ПК\s*\d+[+,\.]?\d*', re.I)
NUM_RANGE_RE = re.compile(r'(\d+[+,\.]?\d*)\s*[-–]\s*(\d+[+,\.]?\d*)')


def norm_text(v):
    if v is None:
        return ''
    return str(v).strip()


def dec(v):
    if v in (None, ''):
        return None
    try:
        return Decimal(str(v).replace(' ', '').replace(',', '.'))
    except Exception:
        return None


def slug(s):
    s = norm_text(s).lower()
    s = re.sub(r'[^a-zа-я0-9]+', '_', s, flags=re.I)
    return s.strip('_')[:60] or 'item'


def parse_pk_value(text):
    t = norm_text(text).replace(' ', '').replace(',', '.').upper().replace('ПК', '')
    if not t:
        return None
    if '+' in t:
        a, b = t.split('+', 1)
        try:
            return Decimal(a) * 100 + Decimal(b)
        except Exception:
            return None
    try:
        return Decimal(t)
    except Exception:
        return None


def parse_range(text):
    txt = norm_text(text)
    if not txt:
        return None, None
    m = PK_RANGE_RE.search(txt)
    if m:
        return parse_pk_value(m.group(1)), parse_pk_value(m.group(2))
    m = NUM_RANGE_RE.search(txt.replace('ПК', ''))
    if m:
        return parse_pk_value(m.group(1)), parse_pk_value(m.group(2))
    hits = PK_ANY_RE.findall(txt)
    if len(hits) >= 2:
        return parse_pk_value(hits[0]), parse_pk_value(hits[1])
    one = parse_pk_value(txt)
    return one, one


def infer_object_type(text):
    low = norm_text(text).lower()
    if 'накопител' in low or 'склад' in low:
        return 'STOCKPILE'
    if 'карьер' in low:
        return 'BORROW_PIT'
    if 'отвал' in low:
        return 'TEMP_DUMP'
    if 'технологическ' in low or 'тех.проезд' in low:
        return 'SERVICE_ROAD'
    if 'притрассов' in low or 'временн' in low or 'подъездн' in low:
        return 'TEMP_ROAD'
    return 'MAIN_TRACK'


def infer_constructive_code(text):
    low = norm_text(text).lower()
    if 'притрассов' in low or 'временн' in low or 'технологическ' in low:
        return 'VPD'
    if 'сва' in low:
        return 'PILE_FIELD'
    if 'иссо' in low or 'площад' in low:
        return 'ISSO_PAD'
    return 'POH'


def work_type_code(work_name):
    h = hashlib.md5(norm_text(work_name).encode('utf-8')).hexdigest()[:8].upper()
    return f'WT_{h}'


def get_id(cur, table, key_col, key_val):
    cur.execute(f"select id from {table} where {key_col}=%s", (key_val,))
    row = cur.fetchone()
    return row[0] if row else None


def get_or_create_section(cur, code, name):
    sid = get_id(cur, 'construction_sections', 'code', code)
    if sid:
        return sid
    cur.execute("insert into construction_sections(code,name,is_active) values (%s,%s,true) returning id", (code, name))
    return cur.fetchone()[0]


def ensure_section_version(cur, section_id, valid_from, pk_start, pk_end, raw_text):
    if pk_start is not None and pk_end is not None and pk_end < pk_start:
        pk_start, pk_end = pk_end, pk_start
    cur.execute("select id from construction_section_versions where section_id=%s and valid_from=%s and pk_start=%s and pk_end=%s", (section_id, valid_from, pk_start, pk_end))
    if cur.fetchone():
        return
    cur.execute("insert into construction_section_versions(section_id,valid_from,pk_start,pk_end,pk_raw_text,is_current,comment) values (%s,%s,%s,%s,%s,true,%s)", (section_id, valid_from, pk_start, pk_end, raw_text, 'Imported from weekly report'))


def get_or_create_work_type(cur, code, name, unit):
    cur.execute("select id from work_types where code=%s", (code,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("insert into work_types(code,name,default_unit,work_group,is_active) values (%s,%s,%s,%s,true) returning id", (code, name[:255], unit or 'шт', 'Импорт отчета'))
    return cur.fetchone()[0]


def get_or_create_object(cur, object_code, name, object_type_code, constructive_code, comment=''):
    cur.execute("select id from objects where object_code=%s", (object_code,))
    row = cur.fetchone()
    if row:
        return row[0]
    object_type_id = get_id(cur, 'object_types', 'code', object_type_code)
    constructive_id = get_id(cur, 'constructives', 'code', constructive_code) if constructive_code else None
    cur.execute("insert into objects(object_code,name,object_type_id,constructive_id,is_active,comment) values (%s,%s,%s,%s,true,%s) returning id", (object_code, name[:255], object_type_id, constructive_id, comment[:500]))
    return cur.fetchone()[0]


def ensure_object_segment(cur, object_id, pk_start, pk_end, raw_text):
    if pk_start is None or pk_end is None:
        return
    if pk_end < pk_start:
        pk_start, pk_end = pk_end, pk_start
    cur.execute("select id from object_segments where object_id=%s and pk_start=%s and pk_end=%s", (object_id, pk_start, pk_end))
    if cur.fetchone():
        return
    cur.execute("insert into object_segments(object_id,pk_start,pk_end,pk_raw_text,comment) values (%s,%s,%s,%s,%s)", (object_id, pk_start, pk_end, raw_text, 'Imported from report'))


def get_or_create_daily_report(cur, report_date, section_id, source_reference, raw_text):
    cur.execute("select id from daily_reports where report_date=%s and section_id=%s and source_reference=%s", (report_date, section_id, source_reference))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("""
        insert into daily_reports(report_date,shift,section_id,source_type,source_reference,raw_text,parse_status,operator_status)
        values (%s,'unknown',%s,%s,%s,%s,'approved','approved') returning id
    """, (report_date, section_id, SOURCE_TYPE, source_reference, raw_text[:1000]))
    return cur.fetchone()[0]


def ensure_project_work_item(cur, object_id, constructive_id, work_type_id, project_volume, unit, source_reference, comment):
    if project_volume is None:
        return None
    cur.execute("select id from project_work_items where object_id=%s and work_type_id=%s and source_reference=%s", (object_id, work_type_id, source_reference))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("insert into project_work_items(object_id,constructive_id,work_type_id,project_volume,unit,source_reference,comment) values (%s,%s,%s,%s,%s,%s,%s) returning id", (object_id, constructive_id, work_type_id, project_volume, unit, source_reference, comment[:500]))
    return cur.fetchone()[0]


def ensure_project_segment(cur, pwi_id, pk_start, pk_end, vol, comment):
    if not pwi_id or pk_start is None or pk_end is None:
        return
    if pk_end < pk_start:
        pk_start, pk_end = pk_end, pk_start
    cur.execute("select id from project_work_item_segments where project_work_item_id=%s and pk_start=%s and pk_end=%s", (pwi_id, pk_start, pk_end))
    if cur.fetchone():
        return
    cur.execute("insert into project_work_item_segments(project_work_item_id,pk_start,pk_end,project_volume_segment,comment) values (%s,%s,%s,%s,%s)", (pwi_id, pk_start, pk_end, vol, comment[:500]))


def ensure_daily_work_item(cur, daily_report_id, report_date, section_id, object_id, constructive_id, work_type_id, work_name_raw, unit, volume, source_reference, comment):
    cur.execute("select id from daily_work_items where daily_report_id=%s and object_id=%s and work_type_id=%s and coalesce(work_name_raw,'')=%s", (daily_report_id, object_id, work_type_id, work_name_raw))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("""
        insert into daily_work_items(daily_report_id,report_date,shift,section_id,object_id,constructive_id,work_type_id,work_name_raw,unit,volume,labor_source_type,contractor_name,comment,approved_by,approved_at)
        values (%s,%s,'unknown',%s,%s,%s,%s,%s,%s,%s,'unknown',null,%s,'agent',now()) returning id
    """, (daily_report_id, report_date, section_id, object_id, constructive_id, work_type_id, work_name_raw[:1000], unit, volume, comment[:1000]))
    return cur.fetchone()[0]


def ensure_daily_segment(cur, dwi_id, pk_start, pk_end, volume_segment, raw_text, comment):
    if not dwi_id or pk_start is None or pk_end is None:
        return
    if pk_end < pk_start:
        pk_start, pk_end = pk_end, pk_start
    cur.execute("select id, volume_segment from daily_work_item_segments where daily_work_item_id=%s and pk_start=%s and pk_end=%s", (dwi_id, pk_start, pk_end))
    row = cur.fetchone()
    if row:
        seg_id, existing_volume = row
        if existing_volume is None and volume_segment is not None:
            cur.execute("update daily_work_item_segments set volume_segment=%s where id=%s", (volume_segment, seg_id))
        return
    cur.execute("insert into daily_work_item_segments(daily_work_item_id,pk_start,pk_end,volume_segment,pk_raw_text,comment) values (%s,%s,%s,%s,%s,%s)", (dwi_id, pk_start, pk_end, volume_segment, raw_text, comment[:500]))


def process_common_sheet(cur, wb, report_date, stats):
    ws = wb[COMMON_SHEET]
    grouped = {}
    current_group = ''
    current_subgroup = ''
    section_bounds = {}
    for section_num in range(1, 9):
        section_name = f'Участок {section_num}'
        section_code = f'UCH_{section_num}'
        section_id = get_or_create_section(cur, section_code, section_name)
        cur.execute("select pk_start, pk_end, pk_raw_text from construction_section_versions where section_id=%s and valid_from=%s order by created_at desc limit 1", (section_id, report_date))
        row = cur.fetchone()
        if row:
            section_bounds[str(section_num)] = row

    for r in range(18, ws.max_row + 1):
        code = norm_text(ws.cell(r, 1).value)
        material_hint = norm_text(ws.cell(r, 2).value)
        section_raw = norm_text(ws.cell(r, 3).value)
        work_name = norm_text(ws.cell(r, 4).value)
        unit = norm_text(ws.cell(r, 6).value)
        project_volume = dec(ws.cell(r, 7).value)
        fact_volume = dec(ws.cell(r, 17).value)
        if not work_name:
            continue
        if not re.fullmatch(r'[1-8]', section_raw):
            if unit == '':
                current_group = work_name
                current_subgroup = ''
            continue
        if unit == '':
            current_subgroup = work_name
            continue
        if fact_volume in (None, Decimal('0')) and project_volume in (None, Decimal('0')):
            continue
        key = (section_raw, material_hint or '0', (current_group or '').strip(), (current_subgroup or '').strip(), work_name.strip(), unit or 'шт')
        item = grouped.setdefault(key, {'project_volume': Decimal('0'), 'fact_volume': Decimal('0'), 'rows': []})
        if project_volume not in (None, Decimal('0')):
            item['project_volume'] += project_volume
        if fact_volume not in (None, Decimal('0')):
            item['fact_volume'] += fact_volume
        item['rows'].append(r)

    common_stats = {'common_grouped_rows': len(grouped), 'common_inserted_daily': 0, 'common_inserted_project': 0, 'common_objects': 0}
    for (section_raw, material_hint, current_group, current_subgroup, work_name, unit), payload in grouped.items():
        section_code = f'UCH_{section_raw}'
        section_name = f'Участок {section_raw}'
        section_id = get_or_create_section(cur, section_code, section_name)
        daily_report_id = get_or_create_daily_report(cur, report_date, section_id, f'{REPORT_PATH.name}:{COMMON_SHEET}:section_{section_raw}', f'Import from {REPORT_PATH.name} / {COMMON_SHEET} / section {section_raw}')
        obj_name_source = current_subgroup or current_group or material_hint or work_name
        object_type_code = infer_object_type((obj_name_source + ' ' + work_name).strip())
        constructive_code = infer_constructive_code((obj_name_source + ' ' + work_name).strip())
        constructive_id = get_id(cur, 'constructives', 'code', constructive_code)
        object_code = f'{section_code}_OBSH_{slug(obj_name_source)}'.upper()[:100]
        object_name = f'{section_name}: {obj_name_source}'[:255]
        obj_existing = get_id(cur, 'objects', 'object_code', object_code)
        object_id = get_or_create_object(cur, object_code, object_name, object_type_code, constructive_code, f'Imported from {COMMON_SHEET}')
        if not obj_existing:
            common_stats['common_objects'] += 1

        if section_raw in section_bounds:
            pk_start, pk_end, pk_raw = section_bounds[section_raw]
            if pk_start is not None and pk_end is not None:
                ensure_object_segment(cur, object_id, pk_start, pk_end, pk_raw or f'By section {section_raw}')

        wt_code = work_type_code(work_name)
        work_type_id = get_or_create_work_type(cur, wt_code, work_name, unit or 'шт')
        src_ref = f"{REPORT_PATH.name}:{COMMON_SHEET}:section_{section_raw}:{slug(obj_name_source)}:{slug(work_name)}"

        if payload['project_volume'] not in (None, Decimal('0')):
            cur.execute("select id from project_work_items where object_id=%s and work_type_id=%s and source_reference=%s", (object_id, work_type_id, src_ref))
            existing = cur.fetchone()
            pwi_id = ensure_project_work_item(cur, object_id, constructive_id, work_type_id, payload['project_volume'], unit or 'шт', src_ref, f'{material_hint} | {current_group} | {current_subgroup}'.strip(' |'))
            if pwi_id and not existing:
                common_stats['common_inserted_project'] += 1

        if payload['fact_volume'] not in (None, Decimal('0')):
            cur.execute("select id from daily_work_items where daily_report_id=%s and work_type_id=%s and coalesce(work_name_raw,'')=%s and unit=%s and volume=%s", (daily_report_id, work_type_id, work_name, unit or 'шт', payload['fact_volume']))
            existing = cur.fetchone()
            dwi_id = ensure_daily_work_item(cur, daily_report_id, report_date, section_id, object_id, constructive_id, work_type_id, work_name, unit or 'шт', payload['fact_volume'], src_ref, f'{material_hint} | {current_group} | {current_subgroup} | rows={payload["rows"]}'.strip(' |'))
            if dwi_id and not existing:
                common_stats['common_inserted_daily'] += 1
            if section_raw in section_bounds:
                pk_start, pk_end, pk_raw = section_bounds[section_raw]
                if pk_start is not None and pk_end is not None:
                    ensure_daily_segment(cur, dwi_id, pk_start, pk_end, payload['fact_volume'], pk_raw or f'By section {section_raw}', 'Imported from common sheet')

    stats.update(common_stats)


def main():
    wb = openpyxl.load_workbook(REPORT_PATH, data_only=True)
    report_date = wb['Аналитика']['E1'].value.date()
    conn = psycopg2.connect(**DB)
    stats = {'sections': 0, 'section_versions': 0, 'objects': 0, 'object_segments': 0, 'work_types': 0, 'daily_reports': 0, 'daily_work_items': 0, 'daily_work_item_segments': 0, 'project_work_items': 0, 'project_segments': 0}
    try:
        with conn:
            with conn.cursor() as cur:
                for sheet_name in SECTION_SHEETS:
                    ws = wb[sheet_name]
                    section_num = re.search(r'(\d+)', sheet_name).group(1)
                    section_code = f'UCH_{section_num}'
                    section_name = f'Участок {section_num}'
                    # locate section boundary row
                    boundary_text = ''
                    for r in range(20, 40):
                        row_vals = [norm_text(ws.cell(r, c).value) for c in range(1, 15)]
                        if any('ПК' in v for v in row_vals):
                            pk_cells = [v for v in row_vals if 'ПК' in v]
                            if pk_cells:
                                boundary_text = pk_cells[-1]
                                break
                    pk_start, pk_end = parse_range(boundary_text)
                    section_id = get_or_create_section(cur, section_code, section_name)
                    stats['sections'] += 1
                    if pk_start is not None and pk_end is not None:
                        before = cur.rowcount
                        ensure_section_version(cur, section_id, report_date, pk_start, pk_end, boundary_text)
                        stats['section_versions'] += 1
                    daily_report_id = get_or_create_daily_report(cur, report_date, section_id, f'{REPORT_PATH.name}:{sheet_name}', f'Import from {REPORT_PATH.name} / {sheet_name}')
                    stats['daily_reports'] += 1

                    current_group = ''
                    for r in range(30, ws.max_row + 1):
                        row_code = norm_text(ws.cell(r, 2).value)
                        material_hint = norm_text(ws.cell(r, 3).value)
                        constructive_or_segment = norm_text(ws.cell(r, 4).value)
                        work_name = norm_text(ws.cell(r, 5).value)
                        unit = norm_text(ws.cell(r, 6).value)
                        project_volume = dec(ws.cell(r, 7).value)
                        daily_volume = dec(ws.cell(r, 15).value)

                        if constructive_or_segment and not work_name:
                            current_group = constructive_or_segment
                            continue
                        if not work_name:
                            continue
                        if daily_volume in (None, Decimal('0')):
                            continue

                        pk_raw_text = constructive_or_segment if ('ПК' in constructive_or_segment or re.search(r'\d+[-–]\d+', constructive_or_segment or '')) else ''
                        object_name_source = current_group or material_hint or work_name
                        object_type_code = infer_object_type(object_name_source + ' ' + work_name)
                        constructive_code = infer_constructive_code(object_name_source + ' ' + work_name)
                        constructive_id = get_id(cur, 'constructives', 'code', constructive_code)

                        obj_slug = slug(object_name_source)
                        object_code = f'{section_code}_{obj_slug}'.upper()[:100]
                        object_name = f'{section_name}: {object_name_source}'[:255]
                        obj_existing = get_id(cur, 'objects', 'object_code', object_code)
                        object_id = get_or_create_object(cur, object_code, object_name, object_type_code, constructive_code, current_group or work_name)
                        if not obj_existing:
                            stats['objects'] += 1

                        opk_start, opk_end = parse_range(pk_raw_text)
                        if opk_start is None and pk_start is not None and pk_end is not None:
                            opk_start, opk_end = pk_start, pk_end
                        if opk_start is not None and opk_end is not None:
                            ensure_object_segment(cur, object_id, opk_start, opk_end, pk_raw_text or boundary_text)
                            stats['object_segments'] += 1

                        wt_code = work_type_code(work_name)
                        wt_existing = get_id(cur, 'work_types', 'code', wt_code)
                        work_type_id = get_or_create_work_type(cur, wt_code, work_name, unit or 'шт')
                        if not wt_existing:
                            stats['work_types'] += 1

                        src_ref = f'{REPORT_PATH.name}:{sheet_name}:row{r}'
                        pwi_existing = None
                        cur.execute("select id from project_work_items where object_id=%s and work_type_id=%s and source_reference=%s", (object_id, work_type_id, src_ref))
                        x = cur.fetchone(); pwi_existing = x[0] if x else None
                        pwi_id = ensure_project_work_item(cur, object_id, constructive_id, work_type_id, project_volume or daily_volume, unit or 'шт', src_ref, current_group or work_name)
                        if pwi_id and not pwi_existing:
                            stats['project_work_items'] += 1
                        if pwi_id and opk_start is not None and opk_end is not None:
                            ensure_project_segment(cur, pwi_id, opk_start, opk_end, project_volume or daily_volume, pk_raw_text or boundary_text)
                            stats['project_segments'] += 1

                        dwi_existing = None
                        cur.execute("select id from daily_work_items where daily_report_id=%s and object_id=%s and work_type_id=%s and coalesce(work_name_raw,'')=%s", (daily_report_id, object_id, work_type_id, work_name))
                        x = cur.fetchone(); dwi_existing = x[0] if x else None
                        dwi_id = ensure_daily_work_item(cur, daily_report_id, report_date, section_id, object_id, constructive_id, work_type_id, work_name, unit or 'шт', daily_volume, src_ref, f'{material_hint} | {current_group}'.strip(' |'))
                        if dwi_id and not dwi_existing:
                            stats['daily_work_items'] += 1
                        if dwi_id and opk_start is not None and opk_end is not None:
                            ensure_daily_segment(cur, dwi_id, opk_start, opk_end, daily_volume, pk_raw_text or boundary_text, 'Imported from section sheet')
                            stats['daily_work_item_segments'] += 1
                process_common_sheet(cur, wb, report_date, stats)
        print(stats)
    finally:
        conn.close()

if __name__ == '__main__':
    main()
