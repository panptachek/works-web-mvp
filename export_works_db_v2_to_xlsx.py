#!/usr/bin/env python3
import os
from pathlib import Path
from decimal import Decimal
from datetime import date, datetime
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DB = dict(
    dbname=os.getenv('DB_NAME', 'works_db_v2'),
    user=os.getenv('DB_USER', 'works_user'),
    password=os.getenv('DB_PASSWORD', ''),
    host=os.getenv('DB_HOST', '127.0.0.1'),
    port=int(os.getenv('DB_PORT', '5433')),
)
OUT = Path('/home/aboba/.openclaw/workspace/works_db_v2_export.xlsx')
TABLES = [
    'construction_sections',
    'construction_section_versions',
    'objects',
    'object_segments',
    'work_types',
    'daily_reports',
    'daily_work_items',
    'daily_work_item_segments',
    'project_work_items',
    'project_work_item_segments',
    'materials',
    'material_movements',
    'report_equipment_units',
]


def py(v):
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.isoformat(sep=' ')
    if isinstance(v, date):
        return v.isoformat()
    return v


def fit_columns(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = '' if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(val) + 2), 60)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    wb = Workbook()
    wb.remove(wb.active)
    conn = psycopg2.connect(**DB)
    try:
        with conn, conn.cursor() as cur:
            for table in TABLES:
                cur.execute(f'select * from {table}')
                rows = cur.fetchall()
                cols = [d.name for d in cur.description]
                ws = wb.create_sheet(title=table[:31])
                ws.append(cols)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                for row in rows:
                    ws.append([py(v) for v in row])
                ws.freeze_panes = 'A2'
                fit_columns(ws)
        wb.save(OUT)
        print(OUT)
    finally:
        conn.close()

if __name__ == '__main__':
    main()
